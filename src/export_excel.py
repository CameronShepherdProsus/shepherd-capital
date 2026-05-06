"""
Excel workbook export.

Generates rebound_capital_fund_tracker.xlsx with 7 tabs.
"""
from __future__ import annotations

import os
from datetime import datetime
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side, numbers,
)
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from src.database import get_session
from src.utils import get_config_value, get_logger

logger = get_logger("export_excel")

# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
GAIN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
LOSS_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")


def _style_header_row(ws, row: int = 1) -> None:
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def _auto_width(ws) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val = str(cell.value or "")
                max_len = max(max_len, len(val))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)


def _write_df(ws, df: pd.DataFrame, start_row: int = 2) -> None:
    """Write DataFrame starting at start_row, with header at start_row-1."""
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=start_row - 1):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)


def _add_table_filter(ws, start_row: int, end_row: int, num_cols: int) -> None:
    if end_row <= start_row:
        return
    last_col = get_column_letter(num_cols)
    ws.auto_filter.ref = f"A{start_row}:{last_col}{end_row}"


# ---------------------------------------------------------------------------
# Load data from DB
# ---------------------------------------------------------------------------

def _load_all() -> dict:
    session = get_session()
    data = {}
    for table, label in [
        ("sources", "sources"),
        ("extracted_recommendations", "recs"),
        ("portfolio_holdings", "holdings"),
        ("fund_performance", "perf"),
        ("prices", "prices"),
        ("ticker_map", "ticker_map"),
    ]:
        try:
            data[label] = pd.read_sql(f"SELECT * FROM {table}", session.bind)
        except Exception:
            data[label] = pd.DataFrame()
    return data


# ---------------------------------------------------------------------------
# Sheet writers
# ---------------------------------------------------------------------------

def _write_summary(ws, data: dict) -> None:
    ws.title = "Summary"
    ts = datetime.now().strftime("%Y-%m-%d %H:%M")

    summary_rows = [
        ["Rebound Capital — Synthetic Fund Tracker", ""],
        [f"Generated: {ts}", ""],
        ["", ""],
        ["Metric", "Value"],
        ["Total scraped articles", len(data["sources"])],
        ["Total extracted ideas", len(data["recs"])],
        ["Active fund holdings", len(data["holdings"])],
    ]

    perf = data["perf"]
    if not perf.empty and "fund_value" in perf.columns:
        latest = perf.sort_values("date").iloc[-1]
        initial = get_config_value("portfolio", "initial_capital", default=100_000)
        summary_rows += [
            ["Current fund value", f"{latest['fund_value']:,.0f}"],
            ["Total return", f"{(latest['fund_value'] - initial) / initial * 100:.2f}%"],
        ]

    holdings = data["holdings"]
    if not holdings.empty and "total_return" in holdings.columns:
        best = holdings.sort_values("total_return", ascending=False).iloc[0]
        worst = holdings.sort_values("total_return").iloc[0]
        summary_rows += [
            ["Best performer", f"{best['ticker']} ({best['total_return']*100:.1f}%)"],
            ["Worst performer", f"{worst['ticker']} ({worst['total_return']*100:.1f}%)"],
        ]

    summary_rows += [
        ["", ""],
        ["Portfolio Methodology", ""],
        ["Weighting", get_config_value("portfolio", "weighting", default="equal")],
        ["Initial capital", get_config_value("portfolio", "initial_capital", default=100_000)],
        ["Rebalance frequency", get_config_value("portfolio", "rebalance_frequency", default="monthly")],
        ["Benchmarks", ", ".join(get_config_value("portfolio", "benchmarks", default=[]))],
        ["Include types", ", ".join(get_config_value("portfolio", "include_types", default=[]))],
        ["Stale threshold (months)", get_config_value("extraction", "stale_months", default=18)],
        ["", ""],
        ["Data source", "Publicly available Rebound Capital content (Substack + website)"],
        ["Compliance note", "Personal research only. No redistribution. Public content only."],
    ]

    for r_idx, row in enumerate(summary_rows, start=1):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            if r_idx == 1:
                cell.font = Font(bold=True, size=14)
            elif row[0] == "Metric" and c_idx == 1:
                cell.fill = SUBHEADER_FILL
                cell.font = Font(bold=True)

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 40
    ws.freeze_panes = "A5"


def _write_holdings(ws, df: pd.DataFrame) -> None:
    ws.title = "Holdings"
    if df.empty:
        ws.cell(1, 1, "No holdings data yet. Run build-portfolio first.")
        return

    cols = [
        "ticker", "company_name", "weight", "entry_date",
        "entry_price", "current_price", "total_return",
        "shares", "market_value", "active_status",
    ]
    available = [c for c in cols if c in df.columns]
    df_out = df[available].copy()

    if "weight" in df_out.columns:
        df_out["weight"] = df_out["weight"].apply(lambda x: f"{x*100:.1f}%" if pd.notna(x) else "")
    if "total_return" in df_out.columns:
        df_out["total_return"] = df_out["total_return"].apply(lambda x: f"{x*100:.2f}%" if pd.notna(x) else "")

    _write_df(ws, df_out)
    _style_header_row(ws, 1)
    _auto_width(ws)
    _add_table_filter(ws, 1, len(df_out) + 1, len(df_out.columns))
    ws.freeze_panes = "A2"

    # Conditional formatting for returns
    for row in ws.iter_rows(min_row=2, max_row=len(df_out) + 1):
        for cell in row:
            if cell.column_letter == "G":  # total_return
                val = str(cell.value or "")
                if val.startswith("-"):
                    cell.fill = LOSS_FILL
                elif val and val != "":
                    cell.fill = GAIN_FILL


def _write_recommendations(ws, df: pd.DataFrame) -> None:
    ws.title = "Recommendations"
    if df.empty:
        ws.cell(1, 1, "No recommendations extracted yet.")
        return

    cols = [
        "ticker", "company_name", "recommendation_type", "confidence",
        "recommendation_date", "thesis_summary", "catalysts", "risks",
        "target_price", "time_horizon", "extraction_method", "excerpt",
    ]
    available = [c for c in cols if c in df.columns]
    df_out = df[available].head(5000)  # cap for Excel sanity

    _write_df(ws, df_out)
    _style_header_row(ws, 1)
    _auto_width(ws)
    _add_table_filter(ws, 1, len(df_out) + 1, len(df_out.columns))
    ws.freeze_panes = "A2"


def _write_sources(ws, df: pd.DataFrame) -> None:
    ws.title = "Source Articles"
    if df.empty:
        ws.cell(1, 1, "No sources scraped yet.")
        return
    cols = ["source_type", "title", "author", "published_date", "url", "status", "fetched_at"]
    available = [c for c in cols if c in df.columns]
    df_out = df[available].copy()

    _write_df(ws, df_out)
    _style_header_row(ws, 1)
    _auto_width(ws)
    _add_table_filter(ws, 1, len(df_out) + 1, len(df_out.columns))
    ws.freeze_panes = "A2"


def _write_price_history(ws, df: pd.DataFrame) -> None:
    ws.title = "Price History"
    if df.empty:
        ws.cell(1, 1, "No price data yet.")
        return
    cols = ["ticker", "date", "open", "high", "low", "close", "adjusted_close", "volume"]
    available = [c for c in cols if c in df.columns]
    # Limit rows for Excel (max ~100k rows for prices)
    df_out = df[available].sort_values(["ticker", "date"], ascending=[True, False]).head(100_000)

    _write_df(ws, df_out)
    _style_header_row(ws, 1)
    _auto_width(ws)
    _add_table_filter(ws, 1, len(df_out) + 1, len(df_out.columns))
    ws.freeze_panes = "A2"


def _write_performance(ws, df: pd.DataFrame) -> None:
    ws.title = "Fund Performance"
    if df.empty:
        ws.cell(1, 1, "No performance data yet.")
        return
    cols = ["date", "fund_value", "daily_return", "cumulative_return",
            "benchmark_spy", "benchmark_qqq", "benchmark_world"]
    available = [c for c in cols if c in df.columns]
    df_out = df[available].copy()

    if "daily_return" in df_out.columns:
        df_out["daily_return"] = df_out["daily_return"].apply(
            lambda x: f"{x*100:.4f}%" if pd.notna(x) else ""
        )
    if "cumulative_return" in df_out.columns:
        df_out["cumulative_return"] = df_out["cumulative_return"].apply(
            lambda x: f"{x*100:.2f}%" if pd.notna(x) else ""
        )

    _write_df(ws, df_out)
    _style_header_row(ws, 1)
    _auto_width(ws)
    ws.freeze_panes = "A2"


def _write_manual_review(ws, recs: pd.DataFrame, ticker_map: pd.DataFrame) -> None:
    ws.title = "Manual Review"

    ws.cell(1, 1, "=== LOW CONFIDENCE TICKER MAPPINGS ===")
    ws.cell(1, 1).font = Font(bold=True, size=12)

    headers = ["company_name", "ticker", "exchange", "confidence", "source", "manual_override"]
    if not ticker_map.empty:
        low_conf = ticker_map[ticker_map.get("confidence", pd.Series(dtype=float)) < 0.7] if "confidence" in ticker_map.columns else ticker_map
        available = [c for c in headers if c in low_conf.columns]
        for c_idx, col in enumerate(available, 1):
            ws.cell(2, c_idx, col).font = Font(bold=True)
        for r_idx, (_, row) in enumerate(low_conf[available].iterrows(), 3):
            for c_idx, col in enumerate(available, 1):
                ws.cell(r_idx, c_idx, row[col])

    row_offset = max(len(ticker_map) + 5, 10) if not ticker_map.empty else 5
    ws.cell(row_offset, 1, "=== LOW CONFIDENCE EXTRACTIONS ===")
    ws.cell(row_offset, 1).font = Font(bold=True, size=12)

    if not recs.empty and "extraction_confidence" in recs.columns:
        low_recs = recs[recs["extraction_confidence"] < 0.5]
        rec_cols = ["ticker", "company_name", "recommendation_type", "confidence", "extraction_method", "excerpt"]
        available = [c for c in rec_cols if c in low_recs.columns]
        for c_idx, col in enumerate(available, 1):
            ws.cell(row_offset + 1, c_idx, col).font = Font(bold=True)
        for r_idx, (_, row) in enumerate(low_recs[available].iterrows(), row_offset + 2):
            for c_idx, col in enumerate(available, 1):
                ws.cell(r_idx, c_idx, row[col])

    _auto_width(ws)


# ---------------------------------------------------------------------------
# Main export function
# ---------------------------------------------------------------------------

def export_excel(output_path: Optional[str] = None) -> Path:
    data = _load_all()

    exports_dir = Path(get_config_value("output", "exports_dir", default="data/exports"))
    exports_dir.mkdir(parents=True, exist_ok=True)

    fname = get_config_value("output", "excel_filename", default="rebound_capital_fund_tracker.xlsx")
    if output_path:
        out_path = Path(output_path)
    else:
        ts = datetime.now().strftime("%Y%m%d_%H%M")
        out_path = exports_dir / fname

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    ws_summary = wb.create_sheet("Summary")
    ws_holdings = wb.create_sheet("Holdings")
    ws_recs = wb.create_sheet("Recommendations")
    ws_sources = wb.create_sheet("Source Articles")
    ws_prices = wb.create_sheet("Price History")
    ws_perf = wb.create_sheet("Fund Performance")
    ws_review = wb.create_sheet("Manual Review")

    _write_summary(ws_summary, data)
    _write_holdings(ws_holdings, data["holdings"])
    _write_recommendations(ws_recs, data["recs"])
    _write_sources(ws_sources, data["sources"])
    _write_price_history(ws_prices, data["prices"])
    _write_performance(ws_perf, data["perf"])
    _write_manual_review(ws_review, data["recs"], data["ticker_map"])

    wb.save(out_path)
    logger.info(f"Excel exported: {out_path}")
    return out_path
